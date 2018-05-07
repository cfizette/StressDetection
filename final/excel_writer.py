from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import datetime
from datetime import datetime as dt


DATA_SHEET = 'Data'
DATE_SHEET = 'Setup'
ENGINE = 'xlsxwriter'
START_DATE_LOC = 'D7'
START_ROW = 3


class XlWriter:
    """Opens and writes patient data to a given excel file.
    Excel file must be derived from project template.

    Attributes:
        filename (string): Path to Excel file to be edited.
        wb (Workbook): Excel workbook.
        ws (Worksheet): Excel worksheet where patient data stored.
        start_date (datetime): Date when recording began.
    
    """
    
    def __init__(self, filename, only_data=False):
        """
        Args:
            filename (string): Path to Excel file to be edited.
        """

        # Remember filename
        self.filename = filename
        
        # Load workbook
        self.wb = load_workbook(filename, data_only=only_data)
        print(type(self.wb))

        # Load worksheet
        
        self.ws = self.wb.get_sheet_by_name(DATA_SHEET)
        self.date_ws = self.wb.get_sheet_by_name(DATE_SHEET)

        # Get start date
        self.start_date = self.date_ws[START_DATE_LOC].value.date()
        print(self.start_date)
        #self.start_date = dt.strptime(self.ws[START_DATE_LOC].value, '%m/%d/%Y').date()


    def row_to_index(self, row):
        """Convert Excel table row number to list index

        Args:
            row (int): Excel row number.
            
        Returns:
            (int): list index of row 
        """
        return row - 1


    def get_hr_column_letter(self, patient):
        """Get HR column letters for patient

        Args:
            patient (int): patient number.
            
        Returns:
            (string) Excel column letter as string
        """
        return get_column_letter(2*patient)


    def get_dfa_column_letter(self, patient):
        """Get DFA column letters for patient
        Args:
            patient (int): patient number.
        Returns:
            (string): Excel column letter as string
        """
        return get_column_letter(2*patient + 1)


    def get_dfa_column(self, patient):
        """Get DFA column object for patient

        Args:
            patient (int): Patient number.

        Returns:
            (:obj: 'list' of :obj: Cell) List of cells in column
        """
        col_letter = self.get_dfa_column_letter(patient)
        return self.ws[col_letter]


    def get_hr_column(self, patient):
        """Get HR column object for patient

        Args:
            patient (int): Patient number.

        Returns:
            (:obj: 'list' of :obj: Cell) List of cells in column
        """
        col_letter = self.get_hr_column_letter(patient)
        return self.ws[col_letter]


    def get_row_from_date(self):
        """Determine what row to write to based on current date.

        Returns:
            (int): Excel row corresponding to current date  
        """
        current_date = datetime.date.today()
        day_delta = current_date - self.start_date
        day_delta_int = day_delta.days
        
##TODO: Handle error
        if day_delta_int < 0:
            pass
            
        return day_delta_int + START_ROW


    def get_index_from_date(self):
        """Determine what list index to write to based on current date.

        Returns:
            (int): List index corresponding to current date
        """
        return self.row_to_index(self.get_row_from_date())


    def save(self):
        """Save changes to workbook"""

        self.wb.save(self.filename)
        print('Data successfully saved')


    def save_patient_data(self, patient, hr, dfa):
        """Write and save patient data.

        Args:
            patient (int): Patient number
            hr (int): HR in bpm
            dfa (float?): DFA measurement
        """
        # Get columns
        hr_col = self.get_hr_column(patient)
        dfa_col = self.get_dfa_column(patient)

        # Get row index
        row_index = self.get_index_from_date()

        # Write data
        hr_col[row_index].value = hr
        dfa_col[row_index].value = dfa

        # Save data
        self.save()










    
